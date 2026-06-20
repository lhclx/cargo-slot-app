# -*- coding: utf-8 -*-
import json, os, datetime, base64, re, hashlib, secrets, time
from flask import Flask, request, jsonify, send_file, session, redirect
from flask_cors import CORS
import openpyxl
import requests as http_requests
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO


def normalize_date(val):
    """标准化日期格式：去掉时间部分，只保留 YYYY-MM-DD"""
    if not val:
        return ''
    s = str(val).strip()
    m = re.match(r'(\d{4}-\d{2}-\d{2})', s)
    return m.group(1) if m else s


def get_week_number(etd_str, eta_str=None):
    """根据ETD/ETA计算自然周编号（ISO周）
    优先用ETD，没有则用ETA
    返回格式: 'W29' (ISO week number)
    """
    date_str = etd_str or eta_str or ''
    if not date_str:
        return ''
    m = re.match(r'(\d{4}-\d{2}-\d{2})', str(date_str))
    if not m:
        return ''
    try:
        d = datetime.datetime.strptime(m.group(1), '%Y-%m-%d')
        return 'W' + str(d.isocalendar()[1])
    except:
        return ''

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'cargo-slot-dev-secret-key-change-in-production')
CORS(app, supports_credentials=True)

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
DATA_FILE = os.path.join(DATA_DIR, 'slots.json')
TRASH_FILE = os.path.join(DATA_DIR, 'trash.json')
USERS_FILE = os.path.join(DATA_DIR, 'users.json')
ORDERS_FILE = os.path.join(DATA_DIR, 'orders.json')

def load_data():
    if not os.path.exists(DATA_FILE):
        return []
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_data(data):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ==================== 用户认证系统 ====================

def load_users():
    if not os.path.exists(USERS_FILE):
        return []
    with open(USERS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_users(users):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

def load_orders():
    """Load orders from JSON file"""
    if not os.path.exists(ORDERS_FILE):
        return []
    with open(ORDERS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_orders(data):
    """Save orders to JSON file"""
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(ORDERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def hash_password(password, salt=None):
    """PBKDF2 SHA256 密码哈希"""
    if salt is None:
        salt = secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 100000)
    return f'pbkdf2:sha256:100000${salt}${dk.hex()}'

def verify_password(password, password_hash):
    if not password_hash or not password_hash.startswith('pbkdf2:'):
        return False
    parts = password_hash.split('$')
    if len(parts) != 3:
        return False
    salt = parts[1]
    return hash_password(password, salt) == password_hash

def get_current_user():
    """从 session 获取当前登录用户"""
    user_id = session.get('user_id')
    if not user_id:
        return None
    for u in load_users():
        if u.get('id') == user_id:
            return u
    return None

def require_auth(fn):
    """装饰器：要求登录"""
    from functools import wraps
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({'error': '未登录', 'need_login': True}), 401
        request.current_user = user
        return fn(*args, **kwargs)
    return wrapper

def require_admin(fn):
    """装饰器：要求管理员权限"""
    from functools import wraps
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({'error': '未登录', 'need_login': True}), 401
        if user.get('role') != 'admin':
            return jsonify({'error': '需要管理员权限'}), 403
        request.current_user = user
        return fn(*args, **kwargs)
    return wrapper

def require_writer(fn):
    """装饰器：要求管理员或操作员（可写）"""
    from functools import wraps
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({'error': '未登录', 'need_login': True}), 401
        if user.get('role') not in ('admin', 'operator'):
            return jsonify({'error': '需要编辑权限'}), 403
        request.current_user = user
        return fn(*args, **kwargs)
    return wrapper

def create_default_admin():
    """如果没有用户，创建默认管理员"""
    users = load_users()
    if not users:
        admin = {
            'id': 1,
            'username': 'admin',
            'password_hash': hash_password('admin123'),
            'role': 'admin',
            'nickname': '系统管理员',
            'permissions': ['all'],
            'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        save_users([admin])

create_default_admin()

# ==================== Initialize Test Orders ====================
def init_test_orders():
    """Create test orders if orders.json is empty"""
    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, 'r', encoding='utf-8') as f:
            existing = json.load(f)
            if existing:
                return  # Already has data
    
    # Ensure we have users to reference
    users = load_users()
    admin_id = 1
    operator_id = 2 if len(users) > 1 else 1
    
    test_orders = [
        {
            'id': 1,
            'job_no': 'JOB001',
            'doc_type': 'JOB',
            'direction': '出口',
            'consignee': 'Test Customer A',
            'customer_ref': 'REF001',
            'po_no': 'PO001',
            'client': 'Client A',
            'shipper': 'Shipper A',
            'biz_type': '海运出口',
            'trade_term': 'FOB',
            'goods_en': 'Electronics',
            'goods_cn': '电子产品',
            'carrier': 'MSC',
            'vessel': 'MSC VESSEL 001',
            'voyage': 'V001',
            'port_loading': 'NINGBO',
            'port_discharge': 'SYDNEY',
            'etd': '2026-07-01',
            'eta': '2026-07-15',
            'container_info': '1x40HQ',
            'container_total': 1,
            'sales': 'admin',
            'operator': 'admin',
            'owner': admin_id,
            'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
            'updated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
            'status': '草稿'
        },
        {
            'id': 2,
            'job_no': 'JOB002',
            'doc_type': 'JOB',
            'direction': '出口',
            'consignee': 'Test Customer B',
            'customer_ref': 'REF002',
            'po_no': 'PO002',
            'client': 'Client B',
            'shipper': 'Shipper B',
            'biz_type': '海运出口',
            'trade_term': 'CIF',
            'goods_en': 'Furniture',
            'goods_cn': '家具',
            'carrier': 'MSK',
            'vessel': 'MSK VESSEL 002',
            'voyage': 'V002',
            'port_loading': 'SHANGHAI',
            'port_discharge': 'MELBOURNE',
            'etd': '2026-07-10',
            'eta': '2026-07-25',
            'container_info': '2x20GP',
            'container_total': 2,
            'sales': 'operator',
            'operator': 'operator',
            'owner': operator_id,
            'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
            'updated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
            'status': '已确认'
        },
        {
            'id': 3,
            'job_no': 'JOB003',
            'doc_type': 'JOB',
            'direction': '进口',
            'consignee': 'Test Customer C',
            'customer_ref': 'REF003',
            'po_no': 'PO003',
            'client': 'Client C',
            'shipper': 'Shipper C',
            'biz_type': '海运进口',
            'trade_term': 'CIF',
            'goods_en': 'Machinery',
            'goods_cn': '机械设备',
            'carrier': 'COSCO',
            'vessel': 'COSCO VESSEL 003',
            'voyage': 'V003',
            'port_loading': 'SYDNEY',
            'port_discharge': 'NINGBO',
            'etd': '2026-07-05',
            'eta': '2026-07-20',
            'container_info': '1x40HQ',
            'container_total': 1,
            'sales': 'admin',
            'operator': 'admin',
            'owner': admin_id,
            'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
            'updated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
            'status': '运输中'
        }
    ]
    save_orders(test_orders)
    print('[Init] Created test orders')

init_test_orders()

# ==================== 认证 API ====================

@app.route('/api/auth/register', methods=['POST'])
@require_admin
def register():
    """管理员注册新用户"""
    body = request.get_json() or {}
    username = str(body.get('username', '')).strip().lower()
    password = str(body.get('password', '')).strip()
    role = str(body.get('role', 'viewer')).strip().lower()
    nickname = str(body.get('nickname', '')).strip()
    permissions = body.get('permissions', [])
    if not username or not password:
        return jsonify({'error': '用户名和密码不能为空'}), 400
    if len(password) < 6:
        return jsonify({'error': '密码至少6位'}), 400
    if role not in ('admin', 'operator', 'viewer'):
        return jsonify({'error': '无效角色'}), 400
    users = load_users()
    if any(u.get('username') == username for u in users):
        return jsonify({'error': '用户名已存在'}), 400
    new_user = {
        'id': max([u['id'] for u in users], default=0) + 1,
        'username': username,
        'password_hash': hash_password(password),
        'role': role,
        'nickname': nickname or username,
        'permissions': permissions if isinstance(permissions, list) else [],
        'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    users.append(new_user)
    save_users(users)
    return jsonify({'ok': True, 'user': {k: v for k, v in new_user.items() if k != 'password_hash'}}), 201

@app.route('/api/auth/login', methods=['POST'])
def login():
    """账号密码登录"""
    body = request.get_json() or {}
    username = str(body.get('username', '')).strip().lower()
    password = str(body.get('password', '')).strip()
    if not username or not password:
        return jsonify({'error': '用户名和密码不能为空'}), 400
    users = load_users()
    user = next((u for u in users if u.get('username') == username), None)
    if not user or not verify_password(password, user.get('password_hash', '')):
        return jsonify({'error': '用户名或密码错误'}), 401
    session['user_id'] = user['id']
    session.permanent = True
    return jsonify({
        'ok': True,
        'user': {
            'id': user['id'],
            'username': user['username'],
            'role': user['role'],
            'nickname': user.get('nickname', ''),
            'permissions': user.get('permissions', [])
        }
    })

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    session.pop('user_id', None)
    return jsonify({'ok': True})

@app.route('/api/auth/me', methods=['GET'])
@require_auth
def get_me():
    user = request.current_user
    return jsonify({
        'id': user['id'],
        'username': user['username'],
        'role': user['role'],
        'nickname': user.get('nickname', ''),
        'permissions': user.get('permissions', [])
    })

@app.route('/api/auth/change-password', methods=['POST'])
@require_auth
def change_password():
    body = request.get_json() or {}
    old_password = str(body.get('old_password', '')).strip()
    new_password = str(body.get('new_password', '')).strip()
    if len(new_password) < 6:
        return jsonify({'error': '新密码至少6位'}), 400
    user = request.current_user
    if not verify_password(old_password, user.get('password_hash', '')):
        return jsonify({'error': '原密码错误'}), 401
    users = load_users()
    for u in users:
        if u['id'] == user['id']:
            u['password_hash'] = hash_password(new_password)
            break
    save_users(users)
    return jsonify({'ok': True})

# ==================== 用户管理 API（管理员）====================

@app.route('/api/users', methods=['GET'])
@require_admin
def get_users():
    users = load_users()
    return jsonify([{
        'id': u['id'],
        'username': u['username'],
        'role': u['role'],
        'nickname': u.get('nickname', ''),
        'permissions': u.get('permissions', []),
        'created_at': u.get('created_at', '')
    } for u in users])

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@require_admin
def update_user(user_id):
    body = request.get_json() or {}
    users = load_users()
    user = next((u for u in users if u['id'] == user_id), None)
    if not user:
        return jsonify({'error': '用户不存在'}), 404
    if 'role' in body:
        role = str(body['role']).strip().lower()
        if role not in ('admin', 'operator', 'viewer'):
            return jsonify({'error': '无效角色'}), 400
        user['role'] = role
    if 'nickname' in body:
        user['nickname'] = str(body['nickname']).strip()
    if 'permissions' in body:
        perms = body['permissions']
        user['permissions'] = perms if isinstance(perms, list) else []
    if 'password' in body and str(body['password']).strip():
        user['password_hash'] = hash_password(str(body['password']).strip())
    save_users(users)
    return jsonify({'ok': True})

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@require_admin
def delete_user(user_id):
    users = load_users()
    if len(users) <= 1:
        return jsonify({'error': '至少需要保留一个用户'}), 400
    new_users = [u for u in users if u['id'] != user_id]
    if len(new_users) == len(users):
        return jsonify({'error': '用户不存在'}), 404
    save_users(new_users)
    return jsonify({'ok': True})

@app.route('/api/slots', methods=['GET'])
@require_auth
def get_slots():
    data = load_data()
    for key in ['status','carrier','operator','port']:
        val = request.args.get(key)
        if val:
            if key == 'port':
                data = [s for s in data if s.get('destination_port') == val or s.get('loading_port') == val]
            else:
                fld = {'status':'status','carrier':'carrier','operator':'operator'}[key]
                data = [s for s in data if s.get(fld) == val]
    etd_start = request.args.get('etd_start','').strip()
    etd_end = request.args.get('etd_end','').strip()
    if etd_start:
        data = [s for s in data if s.get('etd','') >= etd_start]
    if etd_end:
        data = [s for s in data if s.get('etd','') <= etd_end]
    q = request.args.get('q','').lower().strip()
    if q:
        data = [s for s in data if q in str(s.get('so_no','')).lower() or q in str(s.get('booking_ref','')).lower() or q in str(s.get('client','')).lower()]
    week = request.args.get('week','').strip()
    if week:
        data = [s for s in data if get_week_number(s.get('etd'), s.get('eta')) == week]
    # 自动计算周编号
    for s in data:
        s['week'] = get_week_number(s.get('etd'), s.get('eta'))
    return jsonify(data)


@app.route('/api/stats/teu', methods=['GET'])
@require_auth
def stats_teu():
    """按周统计 BSA/LTR/NOR 的 TEU 占用（排除允许超订）"""
    week = request.args.get('week', '').strip()
    data = load_data()
    # 排除允许超订
    data = [s for s in data if '允许超订' not in str(s.get('remark', ''))]
    # 按周筛选
    if week:
        data = [s for s in data if get_week_number(s.get('etd'), s.get('eta')) == week]
    # TEU 换算：20GP=1, 40HQ=2, NOR=2
    teu_map = {'20GP': 1, '40GP': 2, '40HQ': 2, 'NOR': 2}
    # 按 booking_ref 统计
    stats = {}
    for s in data:
        ref = s.get('booking_ref', '其他').strip().upper()
        qty = int(s.get('quantity') or 0)
        ctype = s.get('container_type', '')
        teu = qty * teu_map.get(ctype, 0)
        if ref not in stats:
            stats[ref] = {'teu': 0, '20gp': 0, '40hq': 0, 'nor': 0, 'count': 0}
        stats[ref]['teu'] += teu
        stats[ref]['count'] += qty
        if ctype == '20GP':
            stats[ref]['20gp'] += qty
        elif ctype in ('40HQ', '40GP'):
            stats[ref]['40hq'] += qty
        elif ctype == 'NOR':
            stats[ref]['nor'] += qty
    # 总计
    total_teu = sum(v['teu'] for v in stats.values())
    return jsonify({'week': week or '全部', 'stats': stats, 'total_teu': total_teu})


@app.route('/api/stats/weeks', methods=['GET'])
@require_auth
def stats_weeks():
    """获取所有可用周列表（降序）"""
    data = load_data()
    weeks = set()
    for s in data:
        w = get_week_number(s.get('etd'), s.get('eta'))
        if w:
            weeks.add(w)
    return jsonify(sorted(weeks, reverse=True))

@app.route('/api/slots', methods=['POST'])
@require_writer
def add_slot():
    data = load_data()
    slot = request.get_json()
    slot['id'] = max([s['id'] for s in data], default=0) + 1
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    slot['created_at'] = now
    slot['updated_at'] = now
    data.append(slot)
    save_data(data)
    return jsonify(slot), 201

@app.route('/api/slots/<int:slot_id>', methods=['PUT'])
@require_writer
def update_slot(slot_id):
    data = load_data()
    for i, s in enumerate(data):
        if s['id'] == slot_id:
            upd = request.get_json()
            upd['updated_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            upd['id'] = slot_id
            upd['created_at'] = s.get('created_at', '')
            if 'sale_price' not in upd:
                upd['sale_price'] = s.get('sale_price', '')
            data[i] = upd
            save_data(data)
            return jsonify(upd)
    return jsonify({'error': 'not found'}), 404

@app.route('/api/slots/<int:slot_id>', methods=['DELETE'])
@require_admin
def delete_slot(slot_id):
    """删除舱位到回收站（软删除）"""
    data = load_data()
    slot = None
    for s in data:
        if s['id'] == slot_id:
            slot = s.copy()
            break
    if not slot:
        return jsonify({'error': 'not found'}), 404
    
    # 移到回收站
    trash_data = []
    if os.path.exists(TRASH_FILE):
        with open(TRASH_FILE, 'r', encoding='utf-8') as f:
            trash_data = json.load(f)
    slot['deleted_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    trash_data.append(slot)
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(TRASH_FILE, 'w', encoding='utf-8') as f:
        json.dump(trash_data, f, ensure_ascii=False, indent=2)
    
    # 从主数据中删除
    new_data = [s for s in data if s['id'] != slot_id]
    save_data(new_data)
    return jsonify({'ok': True})

@app.route('/api/trash', methods=['GET'])
@require_auth
def get_trash():
    """获取回收站列表"""
    if not os.path.exists(TRASH_FILE):
        return jsonify({"ok": True, "data": []})
    with open(TRASH_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
        return jsonify({"ok": True, "data": data})

@app.route('/api/trash/<int:slot_id>/restore', methods=['POST'])
@require_admin
def restore_slot(slot_id):
    """恢复舱位从回收站"""
    if not os.path.exists(TRASH_FILE):
        return jsonify({'error': 'trash empty'}), 404
    
    with open(TRASH_FILE, 'r', encoding='utf-8') as f:
        trash_data = json.load(f)
    
    slot = None
    for s in trash_data:
        if s['id'] == slot_id:
            slot = s.copy()
            trash_data.remove(s)
            break
    
    if not slot:
        return jsonify({'error': 'not found in trash'}), 404
    
    # 恢复删除时间
    slot.pop('deleted_at', None)
    
    # 放回主数据
    data = load_data()
    data.append(slot)
    save_data(data)
    
    # 更新回收站
    with open(TRASH_FILE, 'w', encoding='utf-8') as f:
        json.dump(trash_data, f, ensure_ascii=False, indent=2)
    
    return jsonify({'ok': True})

@app.route('/api/trash/<int:slot_id>/permanent', methods=['DELETE'])
@require_admin
def permanent_delete(slot_id):
    """彻底删除（从回收站中移除）"""
    if not os.path.exists(TRASH_FILE):
        return jsonify({'error': 'trash empty'}), 404
    
    with open(TRASH_FILE, 'r', encoding='utf-8') as f:
        trash_data = json.load(f)
    
    original_len = len(trash_data)
    trash_data = [s for s in trash_data if s['id'] != slot_id]
    
    if len(trash_data) == original_len:
        return jsonify({'error': 'not found in trash'}), 404
    
    with open(TRASH_FILE, 'w', encoding='utf-8') as f:
        json.dump(trash_data, f, ensure_ascii=False, indent=2)
    
    return jsonify({'ok': True})

@app.route('/api/trash/clear', methods=['DELETE'])
@require_admin
def clear_trash():
    """清空回收站"""
    if os.path.exists(TRASH_FILE):
        os.remove(TRASH_FILE)
    return jsonify({'ok': True})

COLUMNS = ['ID','船公司','SO号','约号','起运港','目的港','柜型','数量','ETD','ETA','周','使用状态','分配客户','操作归属','销售价格(USD)','备注']

@app.route('/api/slots/export', methods=['GET'])
@require_auth
def export_slots():
    """导出舱位数据为 Excel"""
    data = load_data()
    for key in ['status','carrier','operator','port']:
        val = request.args.get(key)
        if val:
            if key == 'port':
                data = [s for s in data if s.get('destination_port') == val or s.get('loading_port') == val]
            else:
                fld = {'status':'status','carrier':'carrier','operator':'operator'}[key]
                data = [s for s in data if s.get(fld) == val]
    etd_start = request.args.get('etd_start','').strip()
    etd_end = request.args.get('etd_end','').strip()
    if etd_start:
        data = [s for s in data if s.get('etd','') >= etd_start]
    if etd_end:
        data = [s for s in data if s.get('etd','') <= etd_end]
    q = request.args.get('q','').lower().strip()
    if q:
        data = [s for s in data if q in str(s.get('so_no','')).lower() or q in str(s.get('booking_ref','')).lower() or q in str(s.get('client','')).lower()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '舱位列表'

    # 表头
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    hd_font = Font(bold=True, color='FFFFFF', size=11)
    hd_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    hd_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for ci, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = hd_font
        cell.fill = hd_fill
        cell.alignment = hd_align
        cell.border = thin_border

    # 列宽
    widths = [6, 12, 18, 16, 12, 12, 10, 8, 14, 14, 6, 10, 25, 12, 15, 20]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 数据行
    data_align = Alignment(horizontal='center', vertical='center')
    for ri, item in enumerate(sorted(data, key=lambda x: x.get('etd','') or '', reverse=False), 2):
        vals = [
            item.get('id',''),
            item.get('carrier',''),
            item.get('so_no',''),
            item.get('booking_ref',''),
            item.get('loading_port',''),
            item.get('destination_port',''),
            item.get('container_type',''),
            item.get('quantity',1),
            item.get('etd',''),
            item.get('eta',''),
            get_week_number(item.get('etd'), item.get('eta')),
            item.get('status',''),
            item.get('client',''),
            item.get('operator',''),
            item.get('sale_price',''),
            item.get('remark',''),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = data_align
            cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'舱位列表_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    )


@app.route('/api/stats', methods=['GET'])
@require_auth
def get_stats():
    data = load_data()
    stats = {'total': len(data), 'by_status': {}, 'by_carrier': {}, 'by_destination': {}, 'by_operator': {}}
    for s in data:
        for fld, key in [('status','by_status'),('carrier','by_carrier'),('destination_port','by_destination'),('operator','by_operator')]:
            v = s.get(fld, '未知')
            stats[key][v] = stats[key].get(v, 0) + 1
    return jsonify(stats)

@app.route('/api/import', methods=['POST'])
@require_writer
def import_slots():
    """批量导入舱位（支持 Excel/CSV）
    - 按 SO号 匹配：存在则更新，不存在则新增
    - 绝不清空原有数据
    """
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400
    
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': '未选择文件'}), 400
    
    data = load_data()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    imported = 0
    updated = 0
    errors = []
    
    # 建立 SO号 索引（加速查找）
    so_index = {s['so_no']: i for i, s in enumerate(data) if s.get('so_no')}
    
    try:
        # 读取文件
        if file.filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))[1:]  # 跳过表头
        elif file.filename.endswith('.csv'):
            import csv
            content = file.read().decode('utf-8-sig')
            reader = csv.reader(content.splitlines())
            rows = list(reader)[1:]
        else:
            return jsonify({'error': '不支持的文件格式（仅支持 .xlsx / .csv）'}), 400
        
        # 批量导入/更新
        for idx, row in enumerate(rows, start=2):  # Excel 行号从2开始（含表头）
            try:
                if not row[0]:  # 跳过空行
                    continue
                
                so_no = str(row[1] or '').strip()
                if not so_no:
                    errors.append(f'第{idx}行：SO号不能为空')
                    continue
                
                # 构造舱位数据
                slot_data = {
                    'carrier': str(row[0] or ''),
                    'so_no': so_no,
                    'booking_ref': str(row[2] or ''),
                    'loading_port': str(row[3] or '').upper(),
                    'destination_port': str(row[4] or '').upper(),
                    'container_type': str(row[5] or ''),
                    'quantity': int(row[6]) if row[6] else 1,
                    'etd': normalize_date(row[7]),
                    'eta': normalize_date(row[8]),
                    'status': str(row[9] or '空闲'),
                    'client': str(row[10] or ''),
                    'operator': str(row[11] or ''),
                    'sale_price': float(row[12]) if row[12] else '',
                    'remark': str(row[13] or ''),
                    'updated_at': now
                }
                
                # 简单校验
                if not slot_data['carrier']:
                    errors.append(f'第{idx}行：船公司不能为空')
                    continue
                
                # 按 SO号 匹配：存在则更新，不存在则新增
                if so_no in so_index:
                    # 更新现有记录
                    i = so_index[so_no]
                    slot_data['id'] = data[i]['id']
                    slot_data['created_at'] = data[i].get('created_at', now)
                    data[i] = slot_data
                    updated += 1
                else:
                    # 新增记录
                    slot_data['id'] = max([s['id'] for s in data], default=0) + 1
                    slot_data['created_at'] = now
                    data.append(slot_data)
                    # 更新索引
                    so_index[so_no] = len(data) - 1
                    imported += 1
                    
            except Exception as e:
                errors.append(f'第{idx}行：{str(e)}')
        
        save_data(data)
        return jsonify({'imported': imported, 'updated': updated, 'errors': errors, 'total': len(rows)})
    
    except Exception as e:
        return jsonify({'error': f'文件解析失败：{str(e)}'}), 400


@app.route('/api/template')
@require_auth
def download_template():
    """下载导入模板（Excel格式，含下拉菜单）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '舱位导入模板'
    
    # 表头
    headers = ['船公司', 'SO号', '约号', '起运港', '目的港', '柜型', '数量', 'ETD', 'ETA', '使用状态', '分配客户', '操作归属', '销售价格(USD)', '备注']
    ws.append(headers)
    
    # 样式
    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 示例数据
    ws.append(['MSC', 'MSU2894710', 'BK2026001', 'NINGBO', 'SYDNEY', '40HQ', 1, '2026-06-15', '2026-07-01', '空闲', 'MAZEL JM PTY LTD', '张三', 2800.00, '紧急'])
    
    # 列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 20
    
    # ========== 添加下拉菜单 ==========
    
    # 1. 起运港下拉（中国主要港口）
    loading_ports = ['NINGBO', 'SHANGHAI', 'SHENZHEN', 'GUANGZHOU', 'QINGDAO', 'TIANJIN', 'XIAMEN', 'DALIAN', 'SHANTOU', 'FOSHAN']
    dv_loading = DataValidation(type="list", formula1=f'"{",".join(loading_ports)}"', allow_blank=True)
    dv_loading.error = '请选择有效的起运港'
    dv_loading.errorTitle = '起运港错误'
    ws.add_data_validation(dv_loading)
    dv_loading.add(f'D2:D1000')  # 应用到 D 列（起运港）
    
    # 2. 目的港下拉（澳大利亚 + 新西兰主要港口）
    dest_ports = [
        # 澳大利亚
        'SYDNEY', 'MELBOURNE', 'BRISBANE', 'ADELAIDE', 'FREMANTLE', 'DARWIN', 'HOBART',
        # 新西兰
        'AUCKLAND', 'TAURANGA', 'WELLINGTON', 'LYTTELTON', 'TIMARU', 'NAPIER', 'NELSON', 'PICTON'
    ]
    dv_dest = DataValidation(type="list", formula1=f'"{",".join(dest_ports)}"', allow_blank=True)
    dv_dest.error = '请选择有效的目的港'
    dv_dest.errorTitle = '目的港错误'
    ws.add_data_validation(dv_dest)
    dv_dest.add('E2:E1000')  # 应用到 E 列（目的港）
    
    # 3. 柜型下拉
    container_types = ['20GP', '40GP', '40HQ', '45HQ', '20RF', '40RF', 'NOR']
    dv_container = DataValidation(type="list", formula1=f'"{",".join(container_types)}"', allow_blank=True)
    dv_container.error = '请选择有效的柜型'
    dv_container.errorTitle = '柜型错误'
    ws.add_data_validation(dv_container)
    dv_container.add('F2:F1000')  # 应用到 F 列（柜型）
    
    # 4. 使用状态下拉
    statuses = ['空闲', '已预订', '已确认', '已放货', '已完结']
    dv_status = DataValidation(type="list", formula1=f'"{",".join(statuses)}"', allow_blank=True)
    dv_status.error = '请选择有效的使用状态'
    dv_status.errorTitle = '使用状态错误'
    ws.add_data_validation(dv_status)
    dv_status.add('J2:J1000')  # 应用到 J 列（使用状态）
    
    # 5. 船公司下拉（主要船公司）
    carriers = ['MSC', 'MSK', 'CMA CGM', 'COSCO', 'EVERGREEN', 'HAPAG-LLOYD', 'ONE', 'YANG MING', 'ZIM', 'PIL', 'Ocean Network', 'ANL', 'SITC', 'WANHAI', 'RCL', 'FMC', '其他']
    dv_carrier = DataValidation(type="list", formula1=f'"{",".join(carriers)}"', allow_blank=True)
    dv_carrier.error = '请选择有效的船公司'
    dv_carrier.errorTitle = '船公司错误'
    ws.add_data_validation(dv_carrier)
    dv_carrier.add('A2:A1000')  # 应用到 A 列（船公司）
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'舱位导入模板_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ==================== SO 识别功能 ====================

QCLAW_API_BASE = 'http://127.0.0.1:19000/proxy/llm'
QCLAW_API_KEY = os.environ.get('QCLAW_LLM_API_KEY', '')

SO_RECOGNITION_PROMPT = '''You are a professional freight forwarding document recognition assistant. Extract the following information from this Shipping Order (SO) image/PDF. Return ONLY pure JSON (no markdown code block):

Required fields:
- carrier: Shipping line name (abbreviation like MSC, MSK, COSCO, CMA CGM, EVERGREEN, HAPAG-LLOYD, ONE, YANG MING, ZIM, PIL)
- so_no: SO number / Order number
- booking_ref: Booking Reference Number
- loading_port: Port of Loading (code like NINGBO, SHANGHAI, SHENZHEN)
- destination_port: Port of Discharge/Destination (code like SYDNEY, MELBOURNE, BRISBANE, AUCKLAND)
- container_type: Container type (20GP, 40GP, 40HQ, 45HQ)
- quantity: Number of containers (integer)
- etd: ETD - Estimated Time of Departure from FIRST leg/vessel (format YYYY-MM-DD)
- eta: ETA - Estimated Time of Arrival at FINAL destination (format YYYY-MM-DD)

Rules:
1. ETD = first leg departure date
2. ETA = final destination arrival date  
3. If a field is not found, use empty string ""
4. Default quantity to 1 if not found
5. Return ONLY JSON, no other text
6. Port names in UPPERCASE English codes
7. Carrier in standard abbreviation

Example output:
{"carrier":"MSC","so_no":"MSU2894710","booking_ref":"BK2026001","loading_port":"NINGBO","destination_port":"SYDNEY","container_type":"40HQ","quantity":1,"etd":"2026-06-15","eta":"2026-07-10"}
'''

@app.route('/api/recognize-so', methods=['POST'])
@require_writer
def recognize_so():
    """Recognize SO image/PDF and extract slot info"""
    print('[SO识别] 收到请求', flush=True)
    
    if 'file' not in request.files:
        print('[SO识别] 错误：未上传文件', flush=True)
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if not file.filename:
        print('[SO识别] 错误：未选择文件', flush=True)
        return jsonify({'error': 'No file selected'}), 400
    
    print(f'[SO识别] 文件名：{file.filename}，大小：{len(file.read())} bytes', flush=True)
    file.seek(0)  # Reset file pointer
    
    allowed_ext = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.pdf'}
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed_ext:
        print(f'[SO识别] 错误：不支持的格式 {ext}', flush=True)
        return jsonify({'error': 'Unsupported format (png/jpg/jpeg/gif/bmp/webp/pdf)'}), 400
    
    try:
        file_bytes = file.read()
        print(f'[SO识别] 文件读取成功，{len(file_bytes)} bytes，格式：{ext}', flush=True)
        
        # PDF -> image conversion
        if ext == '.pdf':
            print('[SO识别] 开始转换 PDF -> 图片...', flush=True)
            try:
                import pdfplumber
                from PIL import Image as PILImage
                import io
                images = []
                with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                    for page in pdf.pages:
                        im = page.to_image(resolution=150)  # 降低DPI减小文件大小
                        pil_img = im.original  # 获取PIL Image对象
                        # 转换为RGB模式（JPEG不支持RGBA/P模式）
                        if pil_img.mode in ('P', 'PA', 'RGBA', 'LA'):
                            pil_img = pil_img.convert('RGB')
                        buf = io.BytesIO()
                        pil_img.save(buf, format='JPEG', quality=85)  # 用JPEG压缩
                        images.append(buf.getvalue())
                if not images:
                    return jsonify({'error': 'PDF cannot be read'}), 400
                file_bytes = images[0]
                print(f'[SO识别] PDF转图片完成，大小: {len(file_bytes)/1024:.1f}KB', flush=True)
            except ImportError:
                return jsonify({'error': 'PDF support requires pdfplumber: pip install pdfplumber'}), 500
        
        b64 = base64.b64encode(file_bytes).decode('utf-8')
        # 根据来源确定MIME：PDF转的用JPEG，直接上传的用原格式
        if ext == '.pdf':
            mime = 'image/jpeg'
        else:
            mime = 'image/' + ext.lstrip('.')
        data_url = 'data:' + mime + ';base64,' + b64
        
        headers = {
            'Content-Type': 'application/json',
            'Authorization': 'Bearer ' + QCLAW_API_KEY
        }
        payload = {
            'model': 'modelroute',
            'messages': [{
                'role': 'user',
                'content': [
                    {'type': 'text', 'text': SO_RECOGNITION_PROMPT},
                    {'type': 'image_url', 'image_url': {'url': data_url}}
                ]
            }],
            'max_tokens': 1000,
            'temperature': 0.1
        }
        # 自动重试机制（最多3次，递增等待：5秒→10秒→15秒）
        import time as _time
        resp = None
        last_error = ''
        wait_times = [5, 10, 15]  # 递增等待时间
        for attempt in range(3):
            try:
                print(f'[SO识别] API 请求第{attempt+1}次...', flush=True)
                resp = http_requests.post(
                    QCLAW_API_BASE + '/chat/completions',
                    headers=headers, json=payload, timeout=90
                )
                print(f'[SO识别] 状态码: {resp.status_code}', flush=True)
                if resp.status_code == 200:
                    break
                last_error = f'HTTP {resp.status_code}: {resp.text[:100]}'
            except Exception as api_err:
                last_error = str(api_err)[:150]
                print(f'[SO识别] 请求异常: {last_error}', flush=True)
                resp = None
            if attempt < 2:
                wait_sec = wait_times[attempt]
                print(f'[SO识别] 等待{wait_sec}秒后重试...', flush=True)
                _time.sleep(wait_sec)
        
        if not resp or resp.status_code != 200:
            return jsonify({'error': 'AI 请求失败 (已重试3次): ' + last_error}), 502
        result = resp.json()
        content_text = result['choices'][0]['message']['content'].strip()
        if content_text.startswith('```'):
            lines = content_text.split('\n')
            lines = [l for l in lines if not l.strip().startswith('```')]
            content_text = '\n'.join(lines).strip()
        so_data = json.loads(content_text)
        so_data['carrier'] = str(so_data.get('carrier', '')).strip().upper()
        so_data['so_no'] = str(so_data.get('so_no', '')).strip()
        so_data['booking_ref'] = str(so_data.get('booking_ref', '')).strip()
        so_data['loading_port'] = str(so_data.get('loading_port', '')).strip().upper()
        so_data['destination_port'] = str(so_data.get('destination_port', '')).strip().upper()
        so_data['container_type'] = str(so_data.get('container_type', '')).strip().upper()
        so_data['quantity'] = int(so_data.get('quantity', 1) or 1)
        so_data['etd'] = str(so_data.get('etd', '')).strip()
        so_data['eta'] = str(so_data.get('eta', '')).strip()
        return jsonify({'ok': True, 'data': so_data})
    except json.JSONDecodeError as e:
        return jsonify({'error': 'AI response parse failed: ' + str(e), 'raw': content_text[:500]}), 502
    except Exception as e:
        return jsonify({'error': 'SO recognition failed: ' + str(e)}), 500


# ==================== Orders CRUD ====================

# All supported order fields - 新字段列表（2026-06更新）
ORDER_FIELDS = [
    'owner',  # user_id of creator
    'carrier',        # 船司
    'so_no',          # SO号
    'booking_no',     # 约号
    'port_loading',   # 起运港
    'port_discharge', # 目的港
    'container_type',  # 柜型
    'container_seal', # 箱号封号（格式：箱号/封号）
    'client',         # 分配客户
    'sale_price',     # 销售价格
    'remark',         # 备注
    'cut_vgm',        # 截VGM时间（日期）
    'cut_si',         # 截SI时间（日期）
    'truck',          # 拖车（文本）
    'customs',        # 报关（文本）
    'payable',        # 应付（金额）
    'receivable',     # 应收（金额）
    'released',       # 放单（是/否）
]

# Logs API
LOGS_FILE = os.path.join(DATA_DIR, 'logs.json')

def load_logs():
    if not os.path.exists(LOGS_FILE):
        return []
    with open(LOGS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_logs(data):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(LOGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

@app.route('/api/logs/<int:order_id>', methods=['GET'])
@require_auth
def get_order_logs(order_id):
    logs = load_logs()
    order_logs = [l for l in logs if l.get('order_id') == order_id]
    return jsonify(order_logs)

@app.route('/api/logs', methods=['POST'])
@require_writer
def add_order_log():
    logs = load_logs()
    log_entry = request.get_json()
    log_entry['id'] = max([l['id'] for l in logs], default=0) + 1 if logs else 1
    log_entry['time'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_entry['ip'] = request.remote_addr
    logs.append(log_entry)
    save_logs(logs)
    return jsonify(log_entry), 201

@app.route('/api/orders', methods=['GET'])
@require_auth
def get_orders():
    """Get orders with scope filter and pagination"""
    user = request.current_user
    data = load_orders()
    
    # Scope filter: own (default) or all
    scope = request.args.get('scope', 'own').strip()
    
    # Viewer can never see 'all'
    if user.get('role') == 'viewer':
        scope = 'own'
    
    if scope == 'own':
        # Everyone sees only own orders when scope='own'
        data = [o for o in data if str(o.get('owner', '')) == str(user['id'])]
    # Admin with scope=all can see all orders (no filter)
    
    # Add owner names
    try:
        _users = load_users()
        _user_map = {str(u['id']): u.get('nickname', u['username']) for u in _users}
    except:
        _user_map = {}
    for o in data:
        o['_owner_name'] = _user_map.get(str(o.get('owner', '')), 'User#'+str(o.get('owner','')))
    
    # Server-side filtering support
    for field in ORDER_FIELDS:
        val = request.args.get(field, '').strip()
        if val:
            data = [o for o in data if str(o.get(field, '')).lower().find(val.lower()) >= 0]
    
    # Date range filter
    date_field = request.args.get('date_field', 'created_at')
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    if date_from:
        data = [o for o in data if (o.get(date_field, '') or '')[:10] >= date_from]
    if date_to:
        data = [o for o in data if (o.get(date_field, '') or '')[:10] <= date_to]
    
    # Sort by created_at descending
    data.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    
    # Pagination
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    total = len(data)
    start = (page - 1) * per_page
    paged_data = data[start:start + per_page]
    
    return jsonify({'data': paged_data, 'total': total, 'page': page, 'per_page': per_page})

@app.route('/api/orders/<int:order_id>', methods=['GET'])
@require_auth
def get_order(order_id):
    """Get single order by id"""
    data = load_orders()
    for o in data:
        if o['id'] == order_id:
            # Add owner name
            try:
                _users = load_users()
                _u = next((u for u in _users if u['id'] == o.get('owner')), None)
                o['_owner_name'] = _u.get('nickname', _u['username']) if _u else 'Unknown'
            except:
                o['_owner_name'] = 'Unknown'
            return jsonify(o)
    return jsonify({'error': 'not found'}), 404

@app.route('/api/orders', methods=['POST'])
@require_writer
def add_order():
    """Create new order (owner = current user)"""
    user = request.current_user
    data = load_orders()
    order = request.get_json()
    # Sanitize: only keep known fields
    clean = {k: v for k, v in order.items() if k in ORDER_FIELDS}
    clean['id'] = max([o['id'] for o in data], default=0) + 1
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    clean['created_at'] = now
    clean['updated_at'] = now
    clean['owner'] = user['id']  # Set owner to current user
    data.append(clean)
    save_orders(data)
    return jsonify(clean), 201

@app.route('/api/orders/<int:order_id>', methods=['PUT'])
@require_writer
def update_order(order_id):
    """Update order - admin can edit all, operator can only edit own"""
    user = request.current_user
    data = load_orders()
    for i, o in enumerate(data):
        if o['id'] == order_id:
            # Permission check
            if user.get('role') != 'admin' and str(o.get('owner', '')) != str(user['id']):
                return jsonify({'error': 'Permission denied: can only edit own orders'}), 403
            
            upd = request.get_json()
            # Don't allow changing owner
            upd.pop('owner', None)
            clean = {k: v for k, v in upd.items() if k in ORDER_FIELDS or k == 'id' or k == 'created_at' or k == 'updated_at'}
            clean['updated_at'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            clean['id'] = order_id
            clean['created_at'] = o.get('created_at', '')
            clean['owner'] = o.get('owner', user['id'])  # Preserve original owner
            data[i] = clean
            save_orders(data)
            return jsonify(clean)
    return jsonify({'error': 'not found'}), 404

@app.route('/api/orders/<int:order_id>', methods=['DELETE'])
@require_auth
def delete_order(order_id):
    """Delete order - only admin or owner can delete"""
    user = request.current_user
    data = load_orders()
    order = None
    for o in data:
        if o['id'] == order_id:
            order = o
            break
    if not order:
        return jsonify({'error': 'not found'}), 404
    
    # Permission check: admin or owner
    if user.get('role') != 'admin' and str(order.get('owner', '')) != str(user['id']):
        return jsonify({'error': 'Permission denied: only admin or owner can delete'}), 403
    
    new_data = [o for o in data if o['id'] != order_id]
    save_orders(new_data)
    return jsonify({'ok': True})

# ==================== Order Export/Import ====================

@app.route('/api/orders/export', methods=['GET'])
@require_auth
def export_orders():
    """导出订单为 Excel"""
    user = request.current_user
    data = load_orders()
    
    # 权限过滤（与 get_orders 相同逻辑）
    scope = request.args.get('scope', 'own').strip()
    if user.get('role') == 'viewer':
        scope = 'own'
    if scope == 'own':
        data = [o for o in data if str(o.get('owner', '')) == str(user['id'])]
    
    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '订单列表'
    
    # 表头（蓝色背景）
    headers = ['ID', '船司', 'SO号', '约号', '起运港', '目的港', '柜型', '箱号封号', 
               '分配客户', '销售价格', '备注', '截VGM', '截SI', '拖车', '报关', '应付', '应收', '放单']
    
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    hd_font = Font(bold=True, color='FFFFFF', size=11)
    hd_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    hd_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for ci, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = hd_font
        cell.fill = hd_fill
        cell.alignment = hd_align
        cell.border = thin_border
    
    # 列宽
    widths = [8, 12, 15, 15, 12, 12, 10, 18, 15, 12, 20, 12, 12, 15, 15, 12, 12, 10]
    from openpyxl.utils import get_column_letter
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    
    # 数据行
    data_align = Alignment(horizontal='center', vertical='center')
    for ri, item in enumerate(data, 2):
        vals = [
            item.get('id', ''),
            item.get('carrier', ''),
            item.get('so_no', ''),
            item.get('booking_no', ''),
            item.get('port_loading', ''),
            item.get('port_discharge', ''),
            item.get('container_type', ''),
            item.get('container_seal', ''),
            item.get('client', ''),
            item.get('sale_price', ''),
            item.get('remark', ''),
            item.get('cut_vgm', ''),
            item.get('cut_si', ''),
            item.get('truck', ''),
            item.get('customs', ''),
            item.get('payable', ''),
            item.get('receivable', ''),
            '是' if str(item.get('released', '')) == '是' else '否'
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = data_align
            cell.border = thin_border
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'订单导出_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    )


@app.route('/api/orders/import', methods=['POST'])
@require_writer
def import_orders():
    """批量导入订单（按SO号匹配，存在则更新，不存在则新增）"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400
    
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': '未选择文件'}), 400
    
    user = request.current_user
    data = load_orders()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    imported = 0
    updated = 0
    errors = []
    
    # 建立 SO号 索引
    so_index = {str(o.get('so_no', '')): i for i, o in enumerate(data) if o.get('so_no')}
    
    try:
        # 读取文件
        if file.filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))[1:]  # 跳过表头
        else:
            return jsonify({'error': '不支持的文件格式（仅支持 .xlsx）'}), 400
        
        # 批量导入/更新
        for idx, row in enumerate(rows, start=2):
            try:
                if not row[0] and not row[1]:  # 跳过空行
                    continue
                
                so_no = str(row[2] or '').strip() if len(row) > 2 else ''  # 第3列是SO号
                if not so_no:
                    errors.append(f'第{idx}行：SO号不能为空')
                    continue
                
                # 构造订单数据
                order_data = {
                    'carrier': str(row[1] or '').strip() if len(row) > 1 else '',
                    'so_no': so_no,
                    'booking_no': str(row[3] or '').strip() if len(row) > 3 else '',
                    'port_loading': str(row[4] or '').strip() if len(row) > 4 else '',
                    'port_discharge': str(row[5] or '').strip() if len(row) > 5 else '',
                    'container_type': str(row[6] or '').strip() if len(row) > 6 else '',
                    'container_seal': str(row[7] or '').strip() if len(row) > 7 else '',
                    'client': str(row[8] or '').strip() if len(row) > 8 else '',
                    'sale_price': float(row[9]) if len(row) > 9 and row[9] else '',
                    'remark': str(row[10] or '').strip() if len(row) > 10 else '',
                    'cut_vgm': normalize_date(row[11]) if len(row) > 11 else '',
                    'cut_si': normalize_date(row[12]) if len(row) > 12 else '',
                    'truck': str(row[13] or '').strip() if len(row) > 13 else '',
                    'customs': str(row[14] or '').strip() if len(row) > 14 else '',
                    'payable': float(row[15]) if len(row) > 15 and row[15] else '',
                    'receivable': float(row[16]) if len(row) > 16 and row[16] else '',
                    'released': str(row[17] or '').strip() if len(row) > 17 else '否',
                    'updated_at': now
                }
                
                # 按 SO号 匹配：存在则更新，不存在则新增
                if so_no in so_index:
                    # 更新现有记录
                    i = so_index[so_no]
                    order_data['id'] = data[i]['id']
                    order_data['created_at'] = data[i].get('created_at', now)
                    order_data['owner'] = data[i].get('owner', user['id'])
                    data[i] = order_data
                    updated += 1
                else:
                    # 新增记录
                    order_data['id'] = max([o['id'] for o in data], default=0) + 1
                    order_data['created_at'] = now
                    order_data['owner'] = user['id']
                    data.append(order_data)
                    # 更新索引
                    so_index[so_no] = len(data) - 1
                    imported += 1
                    
            except Exception as e:
                errors.append(f'第{idx}行：{str(e)}')
        
        save_orders(data)
        return jsonify({'imported': imported, 'updated': updated, 'errors': errors, 'total': len(rows)})
    
    except Exception as e:
        return jsonify({'error': f'文件解析失败：{str(e)}'}), 400


@app.route('/api/orders/template')
@require_auth
def download_order_template():
    """下载订单导入模板（Excel格式）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '订单导入模板'
    
    # 表头
    headers = ['ID', '船司', 'SO号', '约号', '起运港', '目的港', '柜型', '箱号封号', 
               '分配客户', '销售价格', '备注', '截VGM', '截SI', '拖车', '报关', '应付', '应收', '放单']
    ws.append(headers)
    
    # 样式
    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 示例数据
    ws.append(['', 'MSC', 'MSU2894710', 'BK2026001', 'NINGBO', 'SYDNEY', '40HQ', 'MSCU1234567/123456',
               'Client A', 2800.00, '紧急订单', '2026-07-01', '2026-06-28', '张三卡车', '李四报关', 2000.00, 2800.00, '否'])
    
    # 列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 12
    ws.column_dimensions['Q'].width = 12
    ws.column_dimensions['R'].width = 10
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'订单导入模板_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/orders')
def orders_page():
    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'orders.html'), 'r', encoding='utf-8') as f:
        return f.read(), {'Content-Type': 'text/html; charset=utf-8'}

@app.route('/')
def index():
    return send_html()

def send_html():
    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'index.html'), 'r', encoding='utf-8') as f:
        return f.read(), {'Content-Type': 'text/html; charset=utf-8'}

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    print("=" * 50)
    print(f"  Cargo Slot Management System")
    print(f"  Port: {port}")
    print("=" * 50)
    app.run(host='0.0.0.0', port=port, debug=False)

# For Render.com gunicorn compatibility
app = app
